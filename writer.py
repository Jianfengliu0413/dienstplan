# # writer.py 
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from models import ScheduleModel
import pandas as pd
from statistics import generate_statistics
from report import generate_conflict_report, generate_explanation
from typing import Dict, List, Tuple
from datetime import datetime
from collections import defaultdict
from demand_builder import GLOBAL_STATION

RED_FILL = PatternFill(start_color='EA3323', end_color='EA3323', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='F5C242', end_color='F5C242', fill_type='solid')
BLUE_FILL = PatternFill(start_color='B7C5E4', end_color='B7C5E4', fill_type='solid')
LIGHT_GREEN_FILL = PatternFill(start_color='9FCE63', end_color='9FCE63', fill_type='solid')

def write_output(
    template_path: str,
    output_path: str,
    schedule: ScheduleModel,
    assignment: Dict[int, str],
    duties: List[Tuple[int, str, str]],
    doctors: List[str],
    config: dict,
    solver,
    suggestions_df: pd.DataFrame = None
) -> None:

    
    wb = openpyxl.load_workbook(template_path)
    sheet_name = getattr(schedule, 'sheet_name', None)
    if not sheet_name or sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # --- Build reverse mapping: day_idx -> column ---
    col_for_day = {day_idx: col for col, day_idx in schedule.day_col.items()}

    # Determine the column range for date headers (row 1)
    date_header_row = 1
    start_col = 2
    end_col = ws.max_column
 
    # 构建固定任务集合
    fixed_set = set()
    for doc_name, day_idx, station, abbr in schedule.fixed_assignments:
        fixed_set.add((day_idx, station, abbr))
    # 1. CLEAR ALL WEEKEND CELLS for every doctor row (using header scan)
    for doc_name, row in schedule.doctor_row.items():
        for col in range(start_col, end_col + 1):
            header_cell = ws.cell(row=date_header_row, column=col)
            if isinstance(header_cell.value, datetime) and header_cell.value.weekday() >= 5:
                target_cell = ws.cell(row=row, column=col)
                # Handle merged cells
                if target_cell.coordinate in ws.merged_cells:
                    for merged_range in ws.merged_cells.ranges:
                        if target_cell.coordinate in merged_range:
                            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            try:
                                top_left.value = None
                            except Exception:
                                pass
                            break
                else:
                    try:
                        target_cell.value = None
                    except Exception:
                        pass

    # 2. Clear editable cells (weekdays)
    for row, col in schedule.editable_cells:
        try:
            ws.cell(row=row, column=col).value = None
        except Exception:
            pass

    # 3. Write main assignments (skip weekends, use col_for_day)
    # Load station code mapping (reverse: full name -> code)
    station_code_map_rev = {}
    if 'StationCodeMap' in config:
        df_map = config['StationCodeMap']
        for _, row in df_map.iterrows():
            code = str(row['Code']).strip().lower()
            station = str(row['Station']).strip()
            station_code_map_rev[station] = code

    for i, doc_name in assignment.items():
        day_idx, station, abbr = duties[i]
        if doc_name in schedule.doctor_row:
            row = schedule.doctor_row[doc_name]
            col = col_for_day.get(day_idx)
            if col is None:
                print(f"Warning: day {day_idx} not in col_for_day")
                continue 
            if schedule.days[day_idx].is_weekend and abbr == 'PR':
                code = station_code_map_rev.get(station, station).upper()
                ws.cell(row=row, column=col).value = code
                # 非固定任务 → 红色填充
                if (day_idx, station, abbr) not in fixed_set:
                    ws.cell(row=row, column=col).fill = RED_FILL 
                # 若该单元格不在 editable_cells，打印一次警告（可选）
                # if (row, col) not in schedule.editable_cells:
                #     print(f"Warning: forced write to non-editable weekend cell ({row},{col})")
                continue
            # 其他任务需要 editable
            if (row, col) in schedule.editable_cells:
                try:
                    # 原有写入逻辑（保持不变）
                    if station == GLOBAL_STATION:
                        ws.cell(row=row, column=col).value = abbr
                    else:
                        doc_station = schedule.doctors[doc_name].station
                        if schedule.days[day_idx].is_weekend:
                            code = station_code_map_rev.get(station, station).upper()
                            ws.cell(row=row, column=col).value = code 
                        else:
                            if doc_station != station and abbr in ['ZD', 'SD', 'HD', 'NAZ']:
                                code = station_code_map_rev.get(station, station).upper()
                                ws.cell(row=row, column=col).value = code
                            else:
                                ws.cell(row=row, column=col).value = abbr
                    if abbr == 'ZD':
                        ws.cell(row=row, column=col).fill = BLUE_FILL
                    if abbr == 'SD':
                        ws.cell(row=row, column=col).fill = ORANGE_FILL
                except Exception:
                    pass


    # # 4. Compensatory SD (skip weekends, use col_for_day)
    # add_compensatory_sd(ws, schedule, assignment, duties, doctors, col_for_day)
    # Write fixed assignments (from wishes file)
    for doc_name, day_idx, station, abbr in schedule.fixed_assignments:
        if doc_name in schedule.doctor_row:
            row = schedule.doctor_row[doc_name]
            col = col_for_day.get(day_idx)
            if col is not None:
                try:
                    if abbr == 'PR':
                        # PR on weekend → show station; on weekdays → show "PR"
                        if schedule.days[day_idx].is_weekend:
                            # Write station code (uppercase) instead of full name
                            code = station_code_map_rev.get(station, station).upper()
                            ws.cell(row=row, column=col).value = code
                        else:
                            ws.cell(row=row, column=col).value = abbr
                    else:
                        # For any other duty (NAZ, HD, SD, ZD, KM, SUB, etc.) write the abbreviation
                        ws.cell(row=row, column=col).value = abbr
                except Exception:
                    pass

    # ========== 补偿休息日标记（浅绿色） ==========
    mark_compensatory_days(ws, schedule, assignment, duties, doctors, col_for_day, station_code_map_rev)

    # 5. Add statistics, conflict report, explanation
    if 'Statistics' in wb.sheetnames:
        wb.remove(wb['Statistics'])
    stats_df = generate_statistics(schedule, assignment, duties, doctors)
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a' if 'Statistics' in wb.sheetnames else 'w') as writer:
        stats_df.to_excel(writer, sheet_name='Statistics', index=False)

    if 'ConflictReport' in wb.sheetnames:
        wb.remove(wb['ConflictReport'])
    conflict_df = generate_conflict_report(schedule, assignment, duties, doctors, solver)
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
        conflict_df.to_excel(writer, sheet_name='ConflictReport', index=False)

    if 'Explanation' in wb.sheetnames:
        wb.remove(wb['Explanation'])
    explain_df = generate_explanation(schedule, assignment, duties, doctors)
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
        explain_df.to_excel(writer, sheet_name='Explanation', index=False)

    wb.save(output_path)

    if suggestions_df is not None and not suggestions_df.empty:
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
            suggestions_df.to_excel(writer, sheet_name='DayOffSuggestions', index=False)

def mark_compensatory_days(
    ws,
    schedule: ScheduleModel,
    assignment: Dict[int, str],
    duties: List[Tuple[int, str, str]],
    doctors: List[str],
    col_for_day: Dict[int, int],
    station_code_map_rev: Dict[str, str]
):
    """
    After all assignments are written, mark compensatory days (light green)
    for doctors based on weekend PR (2 PR = 1 day), HD (1 day each), NAZ (1 day each).
    Avoid having more than one doctor from the same station taking comp on the same day,
    and prefer days where the station already has good coverage (fewer vacations/absences).
    """
    from openpyxl.styles import PatternFill
    from collections import defaultdict

    # 1. Count weekend PR, HD, NAZ per doctor
    counts = {doc: {'PR': 0, 'HD': 0, 'NAZ': 0} for doc in doctors}
    for i, doc_name in assignment.items():
        day_idx, station, abbr = duties[i]
        if schedule.days[day_idx].is_weekend:
            if abbr == 'PR':
                counts[doc_name]['PR'] += 1
            elif abbr == 'HD':
                counts[doc_name]['HD'] += 1
            elif abbr == 'NAZ':
                counts[doc_name]['NAZ'] += 1

    # 2. Compute comp needed (round up for PR: 3 PR -> 2 days)
    comp_needed = {}
    for doc in doctors:
        pr = counts[doc]['PR']
        hd = counts[doc]['HD']
        naz = counts[doc]['NAZ']
        comp_needed[doc] = (pr + 1) // 2 + hd + naz

    # 3. Build busy_on_day (doctors already working that day)
    busy_on_day = defaultdict(set)
    for i, doc_name in assignment.items():
        day_idx, _, _ = duties[i]
        busy_on_day[day_idx].add(doc_name)

    # 4. For each doctor, find available weekdays (no duty, not unavailable)
    available = {doc: [] for doc in doctors}
    for day_idx, day in enumerate(schedule.days):
        if day.is_weekend:
            continue
        for doc in doctors:
            if doc in busy_on_day[day_idx]:
                continue
            if (doc, day_idx) in schedule.unavailable:
                continue
            available[doc].append(day_idx)

    # 5. Group doctors by station
    station_doctors = defaultdict(list)
    for doc in doctors:
        station = schedule.doctors[doc].station
        if station is not None:
            station_doctors[station].append(doc)

    # 6. Track how many comps already assigned per station per day (max 1 per station per day)
    station_comp_count = defaultdict(lambda: defaultdict(int))

    # 7. For each station, process doctors with comp needs
    for station, doc_list in station_doctors.items():
        docs_needing = [doc for doc in doc_list if comp_needed.get(doc, 0) > 0]
        # Sort by need descending (more needy first)
        docs_needing.sort(key=lambda d: comp_needed[d], reverse=True)

        for doc in docs_needing:
            needed = comp_needed[doc]
            if needed <= 0:
                continue
            avail_days = available.get(doc, [])
            if not avail_days:
                continue

            # Score each available day: prefer days where the station has the most available doctors
            scored_days = []
            for day_idx in avail_days:
                # Count how many doctors from this station are available (not working, not on vacation)
                station_available = 0
                for other_doc in doc_list:
                    if other_doc == doc:
                        continue
                    if other_doc in busy_on_day[day_idx]:
                        continue
                    if (other_doc, day_idx) in schedule.unavailable:
                        continue
                    station_available += 1
                # Penalize days where this station already has a comp assigned
                existing_comp = station_comp_count[station][day_idx]
                score = station_available - existing_comp * 10  # strong penalty for duplicate comp
                scored_days.append((day_idx, score))

            # Sort by score descending (best coverage first), then by day index
            scored_days.sort(key=lambda x: (-x[1], x[0]))
            selected_days = [day_idx for day_idx, _ in scored_days[:needed]]

            # Assign comp days
            for day_idx in selected_days:
                row = schedule.doctor_row.get(doc)
                if row is None:
                    continue
                col = col_for_day.get(day_idx)
                if col is None:
                    continue
                cell = ws.cell(row=row, column=col)
                if cell.value is None and (row, col) not in schedule.fixed_cells:
                    cell.fill = LIGHT_GREEN_FILL
                    station_comp_count[station][day_idx] += 1

    # 8. Handle doctors without a station (fallback: just take first available days)
    for doc in doctors:
        if schedule.doctors[doc].station is None:
            needed = comp_needed.get(doc, 0)
            if needed <= 0:
                continue
            avail_days = available.get(doc, [])
            if not avail_days:
                continue
            selected = avail_days[:needed]
            for day_idx in selected:
                row = schedule.doctor_row.get(doc)
                if row is None:
                    continue
                col = col_for_day.get(day_idx)
                if col is None:
                    continue
                cell = ws.cell(row=row, column=col)
                if cell.value is None and (row, col) not in schedule.fixed_cells:
                    cell.fill = LIGHT_GREEN_FILL

def add_compensatory_sd(
    ws,
    schedule: ScheduleModel,
    assignment: Dict[int, str],
    duties: List[Tuple[int, str, str]],
    doctors: List[str],
    
    col_for_day: Dict[int, int] 
) -> None:
    """
    Post‑process: for each doctor, count weekend duties and assign the same
    number of SD duties on weekdays where the cell is empty.
    """
    # Build reverse lookup
    assigned = {}
    for i, doc_name in assignment.items():
        day_idx, station, abbr = duties[i]
        assigned[(doc_name, day_idx)] = abbr

    # Count weekend duties per doctor
    weekend_count = {}
    for doc_name in doctors:
        weekend_count[doc_name] = 0

    for i, doc_name in assignment.items():
        day_idx, station, abbr = duties[i]
        if schedule.days[day_idx].is_weekend:
            weekend_count[doc_name] += 1

    # Assign SD on weekdays
    for doc_name in doctors:
        num_sd_needed = weekend_count.get(doc_name, 0)
        if num_sd_needed <= 0:
            continue
        if schedule.doctors[doc_name].fte < 100:
            continue

        weekday_indices = [idx for idx, day in enumerate(schedule.days) if not day.is_weekend]
        assigned_sd = 0
        for day_idx in weekday_indices:
            if assigned_sd >= num_sd_needed:
                break
            if (doc_name, day_idx) in assigned:
                continue
            if (doc_name, day_idx) in schedule.unavailable:
                continue 
            row = schedule.doctor_row.get(doc_name)
            col = col_for_day.get(day_idx)
            if row is None or col is None:
                continue
            if (row, col) not in schedule.editable_cells:
                continue
            if (row, col) in schedule.fixed_cells:
                continue
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            ws.cell(row=row, column=col).value = 'SD'
            assigned_sd += 1
            assigned[(doc_name, day_idx)] = 'SD'

        if assigned_sd < num_sd_needed:
            print(f"Warning: Could not assign all {num_sd_needed} SD duties for {doc_name} (only {assigned_sd} assigned).")


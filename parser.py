# parser.py

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill
from openpyxl.styles.proxy import StyleProxy
from datetime import datetime
import re
import os
import pandas as pd
from models import ScheduleModel, Doctor, Day, DutyType, Station
from demand_builder import GLOBAL_DUTIES, GLOBAL_STATION, MAIN_STATIONS
import shlex
def parse_excluded_list(s: str) -> set:
    if not s:
        return set()
    try:
        return {item.strip() for item in shlex.split(s)}
    except ValueError:
        # fallback to simple comma split if shlex fails
        return {item.strip() for item in s.split(',') if item.strip()}
def is_dark_green(rgb_hex: str, tolerance: int = 10) -> bool:
    """Dark green: G > R and G > B with low brightness."""
    if len(rgb_hex) == 8:
        rgb_hex = rgb_hex[2:]
    if len(rgb_hex) != 6:
        return False
    try:
        r = int(rgb_hex[0:2], 16)
        g = int(rgb_hex[2:4], 16)
        b = int(rgb_hex[4:6], 16)
        # Dark green: G is highest, R and B are low, and overall brightness is low
        return (g > r + tolerance and g > b + tolerance and 
                r < 100 and b < 100 and (r + g + b) < 350)
    except:
        return False

def is_light_green(rgb_hex: str, tolerance: int = 10) -> bool:
    """Light green: G > R and G > B with higher brightness."""
    if len(rgb_hex) == 8:
        rgb_hex = rgb_hex[2:]
    if len(rgb_hex) != 6:
        return False
    try:
        r = int(rgb_hex[0:2], 16)
        g = int(rgb_hex[2:4], 16)
        b = int(rgb_hex[4:6], 16)
        return (g > r + tolerance and g > b + tolerance and 
                (r + g + b) >= 350)
    except:
        return False
def parse_template(template_path: str, config: dict, wishes_path: str = None) -> ScheduleModel:
    """
    Reads the Excel template and builds the ScheduleModel.
    """
    wb = openpyxl.load_workbook(template_path, data_only=True)
    settings = config['Settings'].set_index('Setting')['Value'].to_dict()
    for k, v in settings.items():
        if pd.isna(v):
            settings[k] = ''
        else:
            settings[k] = str(v)

    vacation_color = settings.get('VacationColor', '#00B050')
    # Add unavailable color (default dark green)
    unavailable_color = settings.get('UnavailableColor', '#00B050')
    if len(vacation_color) == 6:
        vacation_color = 'FF' + vacation_color
    if len(unavailable_color) == 6:
        unavailable_color = 'FF' + unavailable_color
    # Normalise unavailable_color to 6-digit hex for exact comparison
    if len(unavailable_color) == 8:
        unavailable_6 = unavailable_color[2:]
    else:
        unavailable_6 = unavailable_color
    sheet_name = settings.get('ScheduleSheet', '')
    if not sheet_name or sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # 获取排除名单
    excluded_doctors_str = settings.get('ExcludedDoctors', '')
    excluded_stations_str = settings.get('ExcludedStations', '')
    excluded_doctors = parse_excluded_list(excluded_doctors_str)
    excluded_stations = parse_excluded_list(excluded_stations_str)
    # --- 1. Parse month and dates ---
    month_cell = ws.cell(row=1, column=1)
    month_val = month_cell.value
    if isinstance(month_val, datetime):
        base_year = month_val.year
        base_month = month_val.month
    elif month_val:
        month_str = str(month_val)
        month_map = {
            'Januar': 1, 'Jan': 1, 'January': 1,
            'Februar': 2, 'Feb': 2, 'February': 2,
            'März': 3, 'Marz': 3, 'March': 3, 'Mar': 3,
            'April': 4, 'Apr': 4,
            'Mai': 5, 'May': 5,
            'Juni': 6, 'June': 6, 'Jun': 6,
            'Juli': 7, 'July': 7, 'Jul': 7,
            'August': 8, 'Aug': 8,
            'September': 9, 'Sep': 9,
            'Oktober': 10, 'October': 10, 'Oct': 10,
            'November': 11, 'Nov': 11,
            'Dezember': 12, 'Dec': 12, 'Dez': 12
        }
        base_month = 10
        base_year = 2026
        for m_name, m_num in month_map.items():
            if m_name.lower() in month_str.lower():
                base_month = m_num
                break
        match = re.search(r'\d{4}', month_str)
        if match:
            base_year = int(match.group())
    else:
        base_month = 10
        base_year = 2026
    print(f"month detected: {base_month}.{base_year}")

    # Parse date headers
    days = []
    day_cols = {}
    start_col = 2
    found_any = False
    for col in range(start_col, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and isinstance(val, str):
            match = re.search(r'(\d{1,2})', val)
            if match:
                day_num = int(match.group(1))
                try:
                    date = datetime(base_year, base_month, day_num)
                    days.append(Day(date=date, weekday=date.strftime('%a'), is_weekend=date.weekday() >= 5))
                    day_cols[col] = len(days) - 1
                    found_any = True
                except ValueError:
                    continue
        elif isinstance(val, datetime):
            day_num = val.day
            try:
                date = datetime(base_year, base_month, day_num)
                days.append(Day(date=date, weekday=date.strftime('%a'), is_weekend=date.weekday() >= 5))
                day_cols[col] = len(days) - 1
                found_any = True
            except ValueError:
                continue

    if not found_any:
        start_col = 3
        for col in range(start_col, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val and isinstance(val, str):
                match = re.search(r'(\d{1,2})', val)
                if match:
                    day_num = int(match.group(1))
                    try:
                        date = datetime(base_year, base_month, day_num)
                        days.append(Day(date=date, weekday=date.strftime('%a'), is_weekend=date.weekday() >= 5))
                        day_cols[col] = len(days) - 1
                        found_any = True
                    except ValueError:
                        continue

    if not days:
        raise ValueError("No valid date headers found in row 1.")

    print(f"Parsed {len(days)} days, first date: {days[0].date.strftime('%Y-%m-%d')}")

    # --- Read SpecialWeekendDays sheet (if present) ---
    special_days_df = config.get('SpecialWeekendDays', pd.DataFrame())
    if not special_days_df.empty:
        # Assume column 'Date' contains dates as strings or Excel dates
        for _, row in special_days_df.iterrows():
            date_val = row['Date']
            if pd.isna(date_val):
                continue
            # Try to parse date
            parsed_date = None
            if isinstance(date_val, (datetime, pd.Timestamp)):
                parsed_date = date_val if isinstance(date_val, datetime) else date_val.to_pydatetime()
            elif isinstance(date_val, str):
                # Try common formats
                for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%m/%d/%Y'):
                    try:
                        parsed_date = datetime.strptime(date_val, fmt)
                        break
                    except ValueError:
                        continue
            elif isinstance(date_val, (int, float)):
                # Excel serial date (assuming 1900 system)
                try:
                    parsed_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_val) - 2)
                except:
                    continue
            if parsed_date:
                # Find matching day in days list
                matched = False
                for day in days:
                    if day.date.year == parsed_date.year and day.date.month == parsed_date.month and day.date.day == parsed_date.day:
                        day.is_weekend = True
                        matched = True
                        print(f"[Special] {day.date.strftime('%Y-%m-%d')} marked as weekend (holiday)")
                        break
                if not matched:
                    print(f"[Warning] Special date {parsed_date.strftime('%Y-%m-%d')} not in current month – ignored.")
                    
    # --- 2. Create the model and set basic attributes ---
    model = ScheduleModel()
    model.days = days
    model.day_col = day_cols
    model.sheet_name = sheet_name

    # Spätdienst IMA row
    ima_sd_days = set()
    for row in range(2, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        if cell_a.value and str(cell_a.value).strip() == 'Spätdienst IMA':
            for col, day_idx in day_cols.items():
                cell = ws.cell(row=row, column=col)
                if cell.value and str(cell.value).strip() == 'SD':
                    ima_sd_days.add(day_idx)
            break

    model.ima_sd_days = ima_sd_days
    print(f"[IMA] SD covered on days: {sorted(ima_sd_days)}")

    # --- 3. Build duty types from config (or fallback) ---
    duty_cfg = config.get('DutyTypes', pd.DataFrame())
    if not duty_cfg.empty:
        for _, row in duty_cfg.iterrows():
            abbr = str(row['Abbr']).strip()
            fullname = str(row.get('FullName', abbr)).strip()
            requires_senior = str(row.get('RequiresSenior', 'No')).strip().upper() == 'YES'
            weekend_only = str(row.get('WeekendOnly', 'No')).strip().upper() == 'YES'
            priority_val = row.get('Priority', 1)
            if pd.isna(priority_val):
                priority_val = 1
            else:
                priority_val = int(priority_val)

            # Read hours from config
            hours_val = row.get('Hours', 8.5)
            if pd.isna(hours_val):
                hours = 8.5
            else:
                hours = float(hours_val)

            dt = DutyType(
                abbr=abbr,
                fullname=fullname,
                requires_senior=requires_senior,
                weekend_only=weekend_only,
                priority=priority_val,
                hours=hours
            )
            model.duty_types[dt.abbr] = dt
    else:
        model.duty_types = {
            'SD': DutyType('SD', 'Spätdienst', False, False, 1),
            'ZD': DutyType('ZD', 'Zwischendienst', False, False, 1),
            'KM': DutyType('KM', 'Knochenmarkentnahme', True, False, 2),
        }
    # Always ensure KM is present
    if 'KM' not in model.duty_types:
        model.duty_types['KM'] = DutyType('KM', 'Knochenmarkentnahme', True, False, 2)

    # --- 4. Load station code mapping ---
    station_code_map = {}
    station_names_set = set()
    if 'StationCodeMap' in config:
        df_map = config['StationCodeMap']
        for _, row in df_map.iterrows():
            code = str(row['Code']).strip().lower()
            station = str(row['Station']).strip()
            station_code_map[code] = station
            station_names_set.add(station)

    # --- 5. Define legend keywords ---
    legend_keywords = {
        'ferien', 'spätdienst ima', 'sd', 'zd', 'km', 'hd', 'naz', 'pr', 'sub',
        'spätdienst', 'zwischendienst', 'knochenmarkentnahme','ambulanzen','Elternzeit'
    }

    # --- 6. Scan rows to identify stations and doctors ---
    station_rows = []      # (row, station_name)
    doctor_rows = []       # (row, name, fte, station_name)
    current_station = None
    found_station_names = set()

    for row in range(2, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        cell_b = ws.cell(row=row, column=2)
        val_a = cell_a.value
        val_b = cell_b.value

        if not val_a and not val_b:
            continue

        name_a = str(val_a).strip() if val_a else ''
        name_b = str(val_b).strip() if val_b else ''

        # Skip legend rows
        if name_a.lower() in legend_keywords or name_b.lower() in legend_keywords:
            continue

        # Check if it's a station header
        is_station = False
        station_name = None

        if name_b and name_b.lower() in station_code_map:
            station_name = station_code_map[name_b.lower()]
            is_station = True
        elif name_a in station_names_set:
            station_name = name_a
            is_station = True

        if is_station:
            current_station = station_name
            found_station_names.add(station_name)
            station_rows.append((row, current_station))
            print(f"\n[Station]: {current_station} ----------- (row {row})")
            continue
        
        # If column A is a known station name, skip (it's a station header, not a doctor)
        if name_a in station_names_set:
            continue

        # Check if it's a doctor row
        if name_a and re.search(r'[a-zA-Z]', name_a):
            if name_a.upper() in model.duty_types:
                continue
            if re.match(r'^\d+(\.\d+)?$', name_a):
                continue

            doctor_name = name_a
            fte = 100

            # Extract FTE from name
            pct_match = re.search(r'(\d{1,3})\s*%', doctor_name)
            if pct_match:
                fte = int(pct_match.group(1))
                doctor_name = re.sub(r'\s*\d{1,3}\s*%', '', doctor_name).strip()
            # Or from column B (e.g., "0.5")
            if name_b:
                try:
                    fte_float = float(name_b)
                    if 0 < fte_float <= 1:
                        fte = int(fte_float * 100)
                    elif 0 < fte_float <= 100:
                        fte = int(fte_float)
                except ValueError:
                    pass
            doctor_name = doctor_name.rstrip('*').strip()

            station = current_station
            if station:
                doctor_rows.append((row, doctor_name, fte, station))
                print(f"[Doctor]: {doctor_name}")
            else:
                doctor_rows.append((row, doctor_name, fte, None))
                print(f"*******[Warning]: Doctor {doctor_name} has no station (row {row})*******")
        else:
            # ignore
            pass
    # --- 7. Add stations to the model ---
    model.found_station_names = found_station_names

    # Create stations from config (if any)
    station_cfg = config.get('Stations', pd.DataFrame())
    for _, row in station_cfg.iterrows():
        name = str(row['Station']).strip()
        req_senior = str(row.get('RequiresSenior', 'No')).strip().upper() == 'YES'
        st = Station(name=name, requires_senior=req_senior)

        weekday_counts_str = str(row.get('WeekdayDutyCounts', ''))
        if weekday_counts_str and weekday_counts_str.strip() and weekday_counts_str.lower() != 'nan':
            for part in weekday_counts_str.split(','):
                if '=' in part:
                    abbr, cnt = part.strip().split('=')
                    st.weekday_demand[abbr.strip()] = int(cnt)

        weekend_counts_str = str(row.get('WeekendDutyCounts', ''))
        if weekend_counts_str and weekend_counts_str.strip() and weekend_counts_str.lower() != 'nan':
            for part in weekend_counts_str.split(','):
                if '=' in part:
                    abbr, cnt = part.strip().split('=')
                    st.weekend_demand[abbr.strip()] = int(cnt)

        st.demand = st.weekday_demand.copy()
        model.stations[name] = st

    # Add stations found in template (if not already in model.stations)
    for station_name in found_station_names:
        if station_name not in model.stations:
            model.stations[station_name] = Station(name=station_name, requires_senior=False)

    # Detect '0' on station rows (no substitution needed)
    for row, station_name in station_rows:
        zero_days = set()
        for col, day_idx in day_cols.items():
            cell = ws.cell(row=row, column=col)
            if cell.value is not None and str(cell.value).strip() == '0':
                zero_days.add(day_idx)
        if station_name in model.stations:
            model.station_zero_days[station_name] = zero_days
    model.stations['Global'] = Station(name='Global', requires_senior=False)
    # --- 8. Define default Active/Weekend based on station ---
    station_defaults = {
        'Sprechstunde PP/Oberärzte': ('No', 'No'),
        'Springer/Konsile/Diagnostik': ('No', 'Yes'),
        '93 (3 IS)': ('No', 'Yes'),
        'Rotation Med I': ('No', 'No'),
        'Rotation': ('No', 'No'),
        'Aufnahme': ('No', 'No'),
        'Forschung': ('No', 'Yes'),
        'Balingen': ('Yes', 'Yes'),
        'Elternzeit': ('No', 'No'),
        "Ambulanzen": ('No', 'No'),
    }
    # Special doctor override
    special_doctor_defaults = {
        'Schwartz': ('No', 'No'),
    }

    # --- 9. Add doctors, handling duplicates (keep first occurrence) ---
    seen_doctors = set()
    model._active_override = {}
    model._weekend_override = {}
    model._inactive_doctors = {}

    # Load the Doctors sheet from config once (already loaded)
    doctors_df = config.get('Doctors', pd.DataFrame())
    # Create a lookup dict for fast access: name -> row (as Series)
    doctor_lookup = {str(row['Name']).strip(): row for _, row in doctors_df.iterrows()} if not doctors_df.empty else {}
 
    for row, name, fte, station in doctor_rows:
        if name in seen_doctors:
            continue
        seen_doctors.add(name)
        # assign category
        if name in doctor_lookup:
            doc_row = doctor_lookup[name]
            category = str(doc_row.get('Category', '')).strip().lower()
            if category:
                # Use explicit category from sheet
                pass
            else:
                # Derive from station
                if station in MAIN_STATIONS:
                    category = 'main'
                elif 'Springer' in station or 'Konsile' in station or 'Diagnostik' in station:
                    category = 'jumper'
                else:
                    category = 'other'
        else:
            # fallback
            category = 'other'
            
        # ----- 1. Determine Active and Weekend from Doctors sheet if available -----
        if name in doctor_lookup:
            doc_row = doctor_lookup[name]
            # Normalise values: "Yes"/"No" or True/False
            active_val = str(doc_row.get('Active', 'Yes')).strip().upper()
            weekend_val = str(doc_row.get('Weekend', 'Yes')).strip().upper()
            # Map to 'Yes'/'No' (handle boolean-like)
            active_default = 'Yes' if active_val in ('YES', 'TRUE', '1') else 'No'
            weekend_default = 'Yes' if weekend_val in ('YES', 'TRUE', '1') else 'No'
        else:
            # Fallback to station_defaults or special overrides
            active_default = 'Yes'
            weekend_default = 'Yes'
            if name in special_doctor_defaults:
                active_default, weekend_default = special_doctor_defaults[name]
            elif station in station_defaults:
                active_default, weekend_default = station_defaults[station]

        # Store overrides (for writing back to the config sheet later)
        model._active_override[name] = active_default
        model._weekend_override[name] = weekend_default

        # ----- 2. Read Allow92KMT and AllowNAZ from Doctors sheet (if present) -----
        allow_92_kmt = False
        allow_naz = False
        if name in doctor_lookup:
            doc_row = doctor_lookup[name]
            if 'Allow92KMT' in doctors_df.columns:
                val = str(doc_row.get('Allow92KMT', 'No')).strip().upper()
                allow_92_kmt = val == 'YES'
            if 'AllowNAZ' in doctors_df.columns:
                val = str(doc_row.get('AllowNAZ', 'No')).strip().upper()
                allow_naz = val == 'YES'
 
        if name in excluded_doctors or (station is not None and station in excluded_stations):
            active_default = 'No'
            print(f"[EXCLUDED] {name} forced inactive (via Settings exclusions)")

        # ----- 3. Add active doctors (skip inactive) -----
        if active_default == 'Yes':
            doc = Doctor(
                name=name,
                fte=fte,
                station=station,
                weekend_available=(weekend_default == 'Yes'),
                allow_92_kmt=allow_92_kmt,
                allow_naz=allow_naz,
                category=category 
            )
            model.doctors[name] = doc
            model.doctor_row[name] = row
        else:
            model._inactive_doctors[name] = {
                'row': row,
                'fte': fte,
                'station': station,
                'active': active_default,
                'weekend': weekend_default,
                'category': category,  # 存储计算好的类别
                'allow_92_kmt': allow_92_kmt,  # 也存一下这两个特殊权限
                'allow_naz': allow_naz
            }
            print(f"[Skipping inactive doctor]: {name} @ {station}")
    # --- 10. Skills ---
    skills_cfg = config.get('Skills', pd.DataFrame())
    for _, row in skills_cfg.iterrows():
        doc_name = row['Doctor']
        duty = row['DutyType']
        if doc_name in model.doctors:
            model.doctors[doc_name].skills.add(duty)

    # Auto-add PR if Weekend=Yes and PR exists
    if 'PR' in model.duty_types:
        for doc in model.doctors.values():
            if doc.weekend_available:
                doc.skills.add('PR')

    # Add SUB for 65 PP doctors
    if 'SUB' not in model.duty_types:
        model.duty_types['SUB'] = DutyType('SUB', 'Substitution', False, False, 1)
    for doc in model.doctors.values():
        if doc.station == '65 PP':
            doc.skills.add('SUB')

    # --- 11. Editable/fixed cells ---
    fixed_vals = set(settings.get('FixedValues', 'x, F, X').split(','))
    fixed_vals = {v.strip() for v in fixed_vals if v.strip()}
    placeholder = settings.get('EditablePlaceholder', '0')
    duty_abbrs = set(model.duty_types.keys())

    for row, name, fte, station in doctor_rows:
        if name not in model.doctors:
            continue 
        for col, day_idx in day_cols.items():
            if day_idx < len(model.days) and model.days[day_idx].is_weekend:
                model.editable_cells.add((row, col))
                if (row, col) in model.fixed_cells:
                    model.fixed_cells.remove((row, col))
                continue
            cell = ws.cell(row=row, column=col)
            val = cell.value
            is_fixed = False
            is_unavailable = False

            # 判断是否为周末
            if day_idx < len(model.days) and model.days[day_idx].is_weekend:
                # 强制可编辑，不检查固定值、颜色等
                model.editable_cells.add((row, col))
                continue  # 跳过后续检查

            if val is not None and str(val).strip() in fixed_vals:
                is_fixed = True
                if str(val).strip() not in duty_abbrs:
                    is_unavailable = True

            # --- Check fill colors ---
            if cell.fill:
                rgb_str = None
                is_dark = False
                is_light = False
                try:
                    if hasattr(cell.fill, 'fgColor'):
                        color = cell.fill.fgColor
                    elif hasattr(cell.fill, 'start_color'):
                        color = cell.fill.start_color
                    else:
                        color = None
                    
                    if color:
                        if color.type == 'theme':
                            theme_idx = color.theme
                            tint = color.tint if color.tint is not None else 0.0
                            # print(f"Cell ({row}, {col}) theme: {theme_idx}, tint: {tint}") # debug
                            # Heuristic: theme 6 is often green (Accent 6)
                            if theme_idx == 6:
                                if tint < -0.1:  # dark green
                                    is_dark = True
                                else:
                                    is_light = True
                        elif color.type == 'rgb' and color.rgb:
                            rgb_str = color.rgb
                            if len(rgb_str) == 8:
                                rgb_str = rgb_str[2:]
                            # print(f"Cell ({row}, {col}) RGB: {rgb_str}") # debug
                            if is_dark_green(rgb_str):
                                is_dark = True
                            elif is_light_green(rgb_str):
                                is_light = True
                except (AttributeError, TypeError):
                    pass

                if is_dark or (rgb_str and is_dark_green(rgb_str)):
                    # print("  -> DARK GREEN detected") # debug
                    is_fixed = True
                    is_unavailable = True
                elif is_light or (rgb_str and is_light_green(rgb_str)):
                    if not hasattr(model, 'soft_unavailable'):
                        model.soft_unavailable = set()
                    model.soft_unavailable.add((name, day_idx))
                        
            if is_fixed:
                model.fixed_cells.add((row, col))
                if is_unavailable:
                    model.unavailable.add((name, day_idx))
            else:
                if val is None or str(val) == placeholder:
                    model.editable_cells.add((row, col))
                else:
                    model.fixed_cells.add((row, col))

    # 加固：遍历所有活跃医生，确保周末列都被添加
    for doc_name, row in model.doctor_row.items():
        for col, day_idx in day_cols.items():
            if day_idx < len(model.days) and model.days[day_idx].is_weekend:
                model.editable_cells.add((row, col))
                if (row, col) in model.fixed_cells:
                    model.fixed_cells.remove((row, col))
    # --- 12. Read HolidayRules sheet ---
    holiday_cfg = config.get('HolidayRules', pd.DataFrame())
    if not holiday_cfg.empty:
        for _, row in holiday_cfg.iterrows():
            doc_name = str(row['Doctor']).strip()
            day_val = row.get('Day')
            if pd.isna(day_val):
                continue
            day_idx = None
            if isinstance(day_val, (int, float)):
                day_num = int(day_val)
                for idx, d in enumerate(model.days):
                    if d.date.day == day_num:
                        day_idx = idx
                        break
            elif isinstance(day_val, (datetime, pd.Timestamp)):
                if isinstance(day_val, pd.Timestamp):
                    day_date = day_val.to_pydatetime()
                else:
                    day_date = day_val
                for idx, d in enumerate(model.days):
                    if d.date.date() == day_date.date():
                        day_idx = idx
                        break
            else:
                try:
                    day_date = pd.to_datetime(day_val)
                    for idx, d in enumerate(model.days):
                        if d.date.date() == day_date.date():
                            day_idx = idx
                            break
                except:
                    pass
            if day_idx is not None:
                if doc_name in model.doctors:
                    model.unavailable.add((doc_name, day_idx))
                    print(f"[HOLIDAY] {doc_name} unavailable on {model.days[day_idx].date}")
                else:
                    print(f"[HOLIDAY WARNING] Doctor '{doc_name}' not found in model.")

    # --- 13. Apply wishes file ---
    if wishes_path and os.path.exists(wishes_path):
        apply_wishes_from_file(model, wishes_path, config, fixed_vals, vacation_color)

    # --- 14. Auto-assign default skill to active doctors with no skills ---
    for doc in model.doctors.values():
        if not doc.skills:
            doc.skills.add('ZD')
            print(f"[INFO] Auto-assigned ZD to {doc.name} (no skills defined)")

    print(f"[DONE!!!]: Parsing complete: found {len(model.doctors)} doctors and {len(model.stations)} stations.")
    return model

def ensure_doctor_active(doc_name: str, model: ScheduleModel) -> bool:
    """如果医生在 _inactive_doctors 中且有固定任务需求，则重新激活他。"""
    if doc_name in model.doctors:
        return True
    if hasattr(model, '_inactive_doctors') and doc_name in model._inactive_doctors:
        info = model._inactive_doctors[doc_name]
        doc = Doctor(
            name=doc_name,
            fte=info['fte'],
            station=info['station'],
            weekend_available=(info['weekend'] == 'Yes'),
            allow_92_kmt=info.get('allow_92_kmt', False),
            allow_naz=info.get('allow_naz', False),
            category=info.get('category', 'other')
        )
        model.doctors[doc_name] = doc
        model.doctor_row[doc_name] = info['row']
        # 从 inactive 列表中移除，防止重复激活
        del model._inactive_doctors[doc_name]
        print(f"[REACTIVATED] {doc_name} due to fixed wish.")
        return True
    return False

def apply_wishes_from_file(model: ScheduleModel, wishes_path: str, config: dict, fixed_vals: set, vacation_color: str):
    """
    从愿望文件读取所有条目，**全部强制转为固定任务**，忽略：
      - 假期/不可用（绿色填充、固定值）
      - 权限（AllowNAZ、Allow92KMT）
      - FTE（全职/兼职）
      - 站匹配（任何医生可去任何站）
      - 技能（自动添加所需技能）
    仅保留同一医生同一天的冲突检测（若已有固定任务，则降级为偏好）。
    """
    wb = openpyxl.load_workbook(wishes_path, data_only=True)
    settings = config['Settings'].set_index('Setting')['Value'].to_dict()
    sheet_name = settings.get('ScheduleSheet', '')
    if not sheet_name or sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # 站代码映射
    station_code_map = {}
    if 'StationCodeMap' in config:
        df_map = config['StationCodeMap']
        for _, row in df_map.iterrows():
            code = str(row['Code']).strip().lower()
            station = str(row['Station']).strip()
            station_code_map[code] = station

    duty_abbrs = set(model.duty_types.keys())

    doctor_name_col = settings.get('DoctorNameColumn', 'A')
    doctor_name_col_idx = column_index_from_string(doctor_name_col)

    # 收集愿望文件中出现的医生（包括 inactive）
    wish_rows = {}
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=doctor_name_col_idx)
        if cell.value and isinstance(cell.value, str):
            name = cell.value.strip().replace('*', '').strip()
            if name in model.doctors or (hasattr(model, '_inactive_doctors') and name in model._inactive_doctors):
                wish_rows[name] = row

    # 解析日期列
    date_row = 1
    day_cols_wish = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=date_row, column=col).value
        if val and isinstance(val, str):
            match = re.search(r'(\d{1,2})', val)
            if match:
                day_num = int(match.group(1))
                for idx, d in enumerate(model.days):
                    if d.date.day == day_num and d.date.month == model.days[0].date.month:
                        day_cols_wish[col] = idx
                        break
        elif isinstance(val, (int, float)):
            day_num = int(val)
            for idx, d in enumerate(model.days):
                if d.date.day == day_num and d.date.month == model.days[0].date.month:
                    day_cols_wish[col] = idx
                    break
        elif isinstance(val, datetime):
            for idx, d in enumerate(model.days):
                if d.date == val:
                    day_cols_wish[col] = idx
                    break

    if not day_cols_wish:
        for col in range(doctor_name_col_idx + 1, ws.max_column + 1):
            day_idx = len(day_cols_wish)
            if day_idx >= len(model.days):
                break
            day_cols_wish[col] = day_idx

    fixed_assignments_set = set()  # 用于检测同一天同一医生的重复固定任务

    for doc_name, row in wish_rows.items():
        for col, day_idx in day_cols_wish.items():
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is None:
                continue
            val_str = str(val).strip()
            if not val_str:
                continue

            # 标准化：去除空格，转换数字
            norm_str = val_str.replace(' ', '')
            if norm_str.endswith('.0'):
                norm_str = norm_str[:-2]
            try:
                numeric_val = float(norm_str)
                if numeric_val.is_integer():
                    norm_str = str(int(numeric_val))
            except ValueError:
                pass
            norm_upper = norm_str.upper()
            norm_lower = norm_str.lower()

            # 激活医生（如果被排除）
            ensure_doctor_active(doc_name, model)
            if doc_name not in model.doctors:
                continue

            # ----- 1. 检查是否为职责缩写 -----
            if norm_upper in duty_abbrs:
                duty_abbr = norm_upper

                # 冲突检测：同一天同一医生已有固定任务
                if (doc_name, day_idx) in fixed_assignments_set:
                    print(f"[CONFLICT] {doc_name} already fixed on {model.days[day_idx].date}, downgrading {duty_abbr} to preference")
                    model.doctors[doc_name].preferences.append((day_idx, duty_abbr, 50))
                    continue
                else:
                    fixed_assignments_set.add((doc_name, day_idx))

                # 确定站（若为全局任务则用 Global）
                if duty_abbr in GLOBAL_DUTIES:
                    station = GLOBAL_STATION
                else:
                    station = model.doctors[doc_name].station
                    if not station:
                        print(f"Warning: {doc_name} has no station, cannot fix {duty_abbr}, will add as preference.")
                        model.doctors[doc_name].preferences.append((day_idx, duty_abbr, 100))
                        continue

                # 自动添加技能（若缺失）
                if duty_abbr not in model.doctors[doc_name].skills:
                    model.doctors[doc_name].skills.add(duty_abbr)
                    print(f"Auto‑added skill {duty_abbr} to {doc_name}")

                model.fixed_assignments.append((doc_name, day_idx, station, duty_abbr))
                model.doctors[doc_name].preferences.append((day_idx, duty_abbr, 100))
                print(f"[FIXED] (duty): {doc_name} -> {duty_abbr} at {station} on {model.days[day_idx].date}")
                continue

            # ----- 2. 检查是否为固定值（如 x, U, dgho 等）→ 记录不可用（但不阻止固定任务）-----
            is_fixed = False
            if val_str in fixed_vals:
                is_fixed = True
            # 绿色填充
            if cell.fill and isinstance(cell.fill, PatternFill):
                color = cell.fill.start_color
                if color and color.rgb:
                    rgb = color.rgb
                    if len(rgb) == 8 and rgb.upper().endswith(vacation_color[-6:].upper()):
                        is_fixed = True
                    elif len(rgb) == 6 and rgb.upper() == vacation_color[-6:].upper():
                        is_fixed = True
            if is_fixed:
                model.unavailable.add((doc_name, day_idx))
                print(f"[unavailable] recorded: {doc_name} on {model.days[day_idx].date} (will be ignored for fixed wishes)")
                continue

            # ----- 3. 检查是否为站代码 -----
            if norm_lower in station_code_map:
                station = station_code_map[norm_lower]

                # 确定班次类型：周末 → PR，否则 → ZD
                duty_abbr = 'PR' if model.days[day_idx].is_weekend else 'ZD'

                # 冲突检测
                if (doc_name, day_idx) in fixed_assignments_set:
                    print(f"[CONFLICT] {doc_name} already fixed on {model.days[day_idx].date}, downgrading {duty_abbr} to preference")
                    model.doctors[doc_name].preferences.append((day_idx, duty_abbr, 50))
                    continue
                else:
                    fixed_assignments_set.add((doc_name, day_idx))

                # 自动添加技能（若缺失）
                if duty_abbr not in model.doctors[doc_name].skills:
                    model.doctors[doc_name].skills.add(duty_abbr)
                    print(f"Auto‑added skill {duty_abbr} to {doc_name}")

                model.fixed_assignments.append((doc_name, day_idx, station, duty_abbr))
                model.doctors[doc_name].preferences.append((day_idx, duty_abbr, 100))
                print(f"[FIXED] (station): {doc_name} -> {duty_abbr} at {station} on {model.days[day_idx].date}")
                continue
 

                # # 冲突检测
                # if (doc_name, day_idx) in fixed_assignments_set:
                #     print(f"[CONFLICT] {doc_name} already fixed on {model.days[day_idx].date}, downgrading {duty_abbr} to preference")
                #     model.doctors[doc_name].preferences.append((day_idx, duty_abbr, 50))
                #     continue
                # else:
                #     fixed_assignments_set.add((doc_name, day_idx))

                # # 自动添加技能
                # if duty_abbr not in model.doctors[doc_name].skills:
                #     model.doctors[doc_name].skills.add(duty_abbr)
                #     print(f"Auto‑added skill {duty_abbr} to {doc_name}")

                # model.fixed_assignments.append((doc_name, day_idx, station, duty_abbr))
                # model.doctors[doc_name].preferences.append((day_idx, duty_abbr, 100))
                # print(f"[FIXED] (station): {doc_name} -> {duty_abbr} at {station} on {model.days[day_idx].date}")
                # continue

            # # 其他任何内容忽略
            # print(f"[Ignored]: {doc_name} on {model.days[day_idx].date} has value '{val_str}'")

    # ----- 解析 NAZ 需求行（如 "NAZ-Dienst(NAZ)"）-----
    naz_demand_days = set()
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        cell_b = ws.cell(row=row, column=2) if ws.max_column >= 2 else None
        val_a = cell_a.value if cell_a else ''
        val_b = cell_b.value if cell_b else ''
        if val_a is None: val_a = ''
        if val_b is None: val_b = ''
        if row in wish_rows.values():
            continue
        if 'NAZ' in str(val_a).upper() or 'NAZ' in str(val_b).upper():
            for col, day_idx in day_cols_wish.items():
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if val is None:
                    continue
                norm_val = str(val).strip().upper()
                if norm_val == 'NAZ':
                    naz_demand_days.add(day_idx)
                    print(f"[NAZ demand] day {day_idx} ({model.days[day_idx].date}) from wishes row {row}")

    model.naz_demand_days = naz_demand_days
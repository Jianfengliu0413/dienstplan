
# scheduler.py
"""
duty scheduler.
"""

import os
import pandas as pd
from config_loader import load_config
from parser import parse_template
from validator import validate
from solver import solve_schedule
from writer import write_output
from openpyxl import load_workbook
from models import ScheduleModel, DutyType
from demand_builder import build_demand
from visualize import visualize_schedule 
import sys
from io import StringIO
from typing import List, Tuple, Dict

def create_default_config(config_path: str):
    """Generate a default Rules.xlsx with sample data."""
    with pd.ExcelWriter(config_path, engine='openpyxl') as writer:
        # Settings
        settings = pd.DataFrame({
            'Setting': ['TemplateFile', 'OutputFile', 'ScheduleSheet', 'ScheduleStartCell', 'ScheduleEndCell',
                        'DoctorNameColumn', 'StationColumn', 'FixedValues', 'EditablePlaceholder', 'VacationColor'],
            'Value': ['Stationsplan_October_2026.xlsx', 'Stationsplan_October_2026_out.xlsx',
                      'Plan', 'B2', '', 'A', 'B', 'x, F, X, U, u, dgho, NAZ, 65P, 92.0, 85.0', '0', '00FF00']
        })
        settings.to_excel(writer, sheet_name='Settings', index=False)

        # Doctors (sample – you will overwrite this later) 
        doctors = pd.DataFrame({
            'Name': ['Dr. Müller', 'Dr. Schmidt', 'Dr. Weber', 'Dr. Fischer'],
            'FTE (%)': [100, 100, 80, 100],
            'Station': ['85 Häm/Onk/Rheu', '92 KMT', '65 PP', '92 KMT'],
            'Active': ['Yes', 'Yes', 'Yes', 'Yes'],
            'Weekend': ['Yes', 'Yes', 'No', 'Yes']   # NEW
        })
        doctors.to_excel(writer, sheet_name='Doctors', index=False)

        # Skills (sample – you will fill this later)
        skills = pd.DataFrame({
            'Doctor': ['Dr. Müller', 'Dr. Schmidt', 'Dr. Weber', 'Dr. Fischer'],
            'DutyType': ['SD', 'ZD', 'DT', 'HD']
        })
        skills.to_excel(writer, sheet_name='Skills', index=False)

        # Stations (sample – you will fill DutyCounts later)
        stations = pd.DataFrame({
            'Station': ['85 Häm/Onk/Rheu', '92 KMT', '65 PP', '65 LAF', 'HD'],
            'RequiresSenior': ['No', 'No', 'No', 'No', 'No'],
            'DutyCounts': ['', '', '', '', '']  # leave empty – you fill them
        })
        stations.to_excel(writer, sheet_name='Stations', index=False)

        # DutyTypes – now includes DT and HD
        duty_types = pd.DataFrame({
            'Abbr': ['SD', 'ZD', 'KM', 'DT', 'HD'],
            'FullName': ['Spätdienst', 'Zwischendienst', 'Knochenmarkentnahme', 'Tagesdienst', 'Halbtagdienst'],
            'RequiresSenior': ['No', 'No', 'Yes', 'No', 'No'],
            'WeekendOnly': ['No', 'No', 'No', 'No', 'No'],
            'Priority': [1, 1, 2, 1, 1]
        })
        duty_types.to_excel(writer, sheet_name='DutyTypes', index=False)

        # GeneralRules
        general = pd.DataFrame({
            'RuleName': ['MaxConsecutiveWorkDays', 'MaxDutiesPerWeek'],
            'Value': [6, 5]
        })
        general.to_excel(writer, sheet_name='GeneralRules', index=False)

        # HolidayRules (vacation)
        holidays = pd.DataFrame(columns=['Doctor', 'Day', 'Type'])
        holidays.to_excel(writer, sheet_name='HolidayRules', index=False)

        # Preferences
        preferences = pd.DataFrame(columns=['Doctor', 'Day', 'DutyType', 'Priority'])
        preferences.to_excel(writer, sheet_name='Preferences', index=False)

        # Penalties
        penalties = pd.DataFrame({
            'Penalty': ['Preference', 'WorkloadBalance', 'WeekendBalance'],
            'Weight': [10, 20, 15]
        })
        penalties.to_excel(writer, sheet_name='Penalties', index=False)

        # Constraints – add WeekendOnlyFullTime
        constraints = pd.DataFrame({
            'Constraint': ['MaxConsecutive', 'MaxPerWeek', 'WeekendOnly', 'SeniorRequired', 'WeekendOnlyFullTime'],
            'Enabled': ['Yes', 'Yes', 'Yes', 'Yes', 'Yes']
        })
        constraints.to_excel(writer, sheet_name='Constraints', index=False)

        # OutputOptions
        output = pd.DataFrame({
            'Option': ['IncludeStatistics', 'IncludeConflictReport', 'IncludeExplanation'],
            'Value': ['Yes', 'Yes', 'Yes']
        })
        output.to_excel(writer, sheet_name='OutputOptions', index=False)
def get_default_station_code_map() -> pd.DataFrame:
    data = [
        ("1", "65 PP"),
        ("2", "65 LAF"),
        ("3", "85 Häm/Onk/Rheu"),
        ("4", "92 KMT"),
        ("5", "64 TK PP"),
        ("6", "64 TK Kasse"),
        ("7", "Sonographie"),
        ("8", "Amb 1 (Lymphom/allgemein)"),
        ("9", "Amb 2 (allgemein/Gerinnung)"),
        ("10", "Amb 3 (spezialisiert mix)"),
        ("11", "Amb 4 (Myelom)"),
        ("12", "KMT 1"),
        ("13", "KMT 2"),
        ("14", "Rheuma 1"),
        ("15", "Rheuma 2"),
        ("16", "Rheuma 3"),
        ("17", "INDIRA"),
        ("17a", "GBA-Rheumazentrum 50 %"),
        ("18", "Leukapherese"),
        ("19", "Sprechstunde PP/Oberärzte"),
        ("20", "Springer/Konsile/Diagnostik"),
        ("21", "Balingen"),
        ("22", "Forschung"),
        ("23", "Labor"),
        ("24", "93 (3 IS)"),
        ("25", "Rotation Med I"),
        ("28", "Aufnahme"),
        ("25-copy", "Rotation"),
        ("92", "92 KMT"),
        ("85", "85 Häm/Onk/Rheu"),
        ("65p", "65 PP"),
        ("65", "65 LAF"),
        ("999", "Elternzeit"),
    ]
    return pd.DataFrame(data, columns=["Code", "Station"])
def write_missing_config_sheets(model: ScheduleModel, config_path: str):
    from openpyxl import load_workbook
    import pandas as pd

    # 1. Load existing Stations data (to preserve user-entered DutyCounts)
    existing_stations = {}
    try:
        df_existing = pd.read_excel(config_path, sheet_name='Stations')
        for _, row in df_existing.iterrows():
            name = str(row['Station']).strip()
            weekday = str(row.get('WeekdayDutyCounts', '')).strip()
            weekend = str(row.get('WeekendDutyCounts', '')).strip()
            if weekday == 'nan':
                weekday = ''
            if weekend == 'nan':
                weekend = ''
            existing_stations[name] = (weekday, weekend)
    except:
        pass

    # 2. Merge with detected stations
    new_stations = {}
    for station_name in sorted(model.found_station_names):
        if station_name in existing_stations:
            weekday, weekend = existing_stations[station_name]
        else:
            weekday = ''
            if station_name in ['65 LAF', '65 PP', '85 Häm/Onk/Rheu', '92 KMT']:
                weekend = 'PR=1'
            else:
                weekend = ''
        new_stations[station_name] = (weekday, weekend)

    # Keep existing stations not in template
    for name, (weekday, weekend) in existing_stations.items():
        if name not in new_stations:
            new_stations[name] = (weekday, weekend)

    stations_data = []
    for name, (weekday, weekend) in sorted(new_stations.items()):
        stations_data.append([name, 'No', weekday, weekend])
    df_stations = pd.DataFrame(stations_data, columns=['Station', 'RequiresSenior', 'WeekdayDutyCounts', 'WeekendDutyCounts'])

    # 3. Prepare Doctors data
    doctors_data = []
    for doc in model.doctors.values():
        name = doc.name
        active = getattr(model, '_active_override', {}).get(name, 'Yes')
        weekend = getattr(model, '_weekend_override', {}).get(name, 'Yes')
        doctors_data.append([name, doc.fte, doc.station, active, weekend])

    if hasattr(model, '_inactive_doctors'):
        for name, info in model._inactive_doctors.items():
            doctors_data.append([
                name,
                info['fte'],
                info['station'],
                info['active'],
                info['weekend']
            ])
    df_doctors = pd.DataFrame(doctors_data, columns=['Name', 'FTE (%)', 'Station', 'Active', 'Weekend'])

    # 4. Load workbook, delete Doctors and Stations if they exist
    wb = load_workbook(config_path)
    for sheet_name in ['Doctors', 'Stations']:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    wb.save(config_path)

    # 5. Write new Doctors and Stations sheets
    with pd.ExcelWriter(config_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        df_doctors.to_excel(writer, sheet_name='Doctors', index=False)
        df_stations.to_excel(writer, sheet_name='Stations', index=False)

    # 6. Create/update other sheets (Skills, DutyTypes, etc.) without deleting
    with pd.ExcelWriter(config_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        # Skills
        if 'Skills' not in wb.sheetnames:
            pd.DataFrame(columns=['Doctor', 'DutyType']).to_excel(writer, sheet_name='Skills', index=False)
        else:
            df_skills = pd.read_excel(config_path, sheet_name='Skills')
            if df_skills.empty:
                pd.DataFrame(columns=['Doctor', 'DutyType']).to_excel(writer, sheet_name='Skills', index=False)

        # DutyTypes – if missing, create default
        if 'DutyTypes' not in wb.sheetnames:
            duty_data = []
            for abbr, dt in model.duty_types.items():
                duty_data.append([abbr, dt.fullname,
                                  'Yes' if dt.requires_senior else 'No',
                                  'Yes' if dt.weekend_only else 'No',
                                  dt.priority])
            df_duty = pd.DataFrame(duty_data, columns=['Abbr', 'FullName', 'RequiresSenior', 'WeekendOnly', 'Priority'])
            df_duty.to_excel(writer, sheet_name='DutyTypes', index=False)

        # --- Create StationCodeMap if missing ---
        if 'StationCodeMap' not in wb.sheetnames:
            df_map = get_default_station_code_map()
            df_map.to_excel(writer, sheet_name='StationCodeMap', index=False)

        # Other required sheets (if missing)
        required_sheets = {
            'GeneralRules': pd.DataFrame({'RuleName': ['MaxConsecutiveWorkDays', 'MaxDutiesPerWeek'], 'Value': [6, 5]}),
            'Penalties': pd.DataFrame({'Penalty': ['Preference', 'WorkloadBalance', 'WeekendBalance', 'CrossStation'], 'Weight': [10, 100, 15, 30]}),
            'Constraints': pd.DataFrame({'Constraint': ['MaxConsecutive', 'MaxPerWeek', 'WeekendOnly', 'SeniorRequired', 'WeekendAvailability', 'WeekendOnlyFullTime', 'WeekendOnlyForSkilled', 'MaxOneWeekendPerDoctor'], 'Enabled': ['Yes', 'Yes', 'No', 'Yes', 'Yes', 'Yes', 'Yes', 'No']}),
            'Preferences': pd.DataFrame(columns=['Doctor', 'Day', 'DutyType', 'Priority']),
            'OutputOptions': pd.DataFrame({'Option': ['IncludeStatistics', 'IncludeConflictReport', 'IncludeExplanation'], 'Value': ['Yes', 'Yes', 'Yes']})
        }
        for sheet_name, df_template in required_sheets.items():
            if sheet_name not in wb.sheetnames:
                df_template.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                df_existing = pd.read_excel(config_path, sheet_name=sheet_name)
                if df_existing.empty:
                    df_template.to_excel(writer, sheet_name=sheet_name, index=False)

    print("Doctors and Stations sheets replaced (no duplicates).")
    print("Other configuration sheets (Skills, DutyTypes, etc.) are ready.")

def write_working_hours(config_path: str, hours: Dict[str, float]):
    df = pd.DataFrame(list(hours.items()), columns=['Doctor', 'Hours'])
    with pd.ExcelWriter(config_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='WorkingHours', index=False)
        
def write_skills_auto(model: ScheduleModel, config_path: str):
    """Write a sheet with all doctors and all possible duty types as skills.""" 
    
    # Get all duty type abbreviations from the model
    duty_abbrs = list(model.duty_types.keys())
    # Add special skill tokens
    special_flags = [] #['Senior', 'Weekend']
    all_skills = duty_abbrs + special_flags
    
    # Build a DataFrame with one row per doctor and one column per duty type
    # Or, simpler: long format: Doctor, DutyType
    data = []
    for doc_name in model.doctors.keys():
        for duty in all_skills:
            data.append([doc_name, duty])
    df = pd.DataFrame(data, columns=['Doctor', 'DutyType'])
    
    # Write to a new sheet 'Skills-Auto' (overwrite if exists)
    with pd.ExcelWriter(config_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='Skills-Auto', index=False)
    
    print("Generated 'Skills-Auto' sheet with all doctors × all duty types.")
    print("   Copy the rows you need into the 'Skills' sheet and delete the rest.")

def generate_day_off_suggestions(
    schedule: ScheduleModel,
    assignment: Dict[int, str],
    duties: List[Tuple[int, str, str]],
    doctors: List[str],
    duty_hours: List[float],
    initial_hours: Dict[str, float],
    final_hours: Dict[str, float]
) -> pd.DataFrame:
    """
    For each part‑time doctor (FTE < 100), suggest removing some blank weekdays
    to bring total hours closer to the FTE‑based target.
    Returns a DataFrame with: Doctor, Current Hours, Target Hours, Excess, Suggested Days.
    """
    rows = []
    total_month_hours = sum(duty_hours)
    total_initial_hours = sum(initial_hours.values())
    total_fte = sum(doc.fte for doc in schedule.doctors.values()) / 100.0
    weekday_indices = [idx for idx, day in enumerate(schedule.days) if not day.is_weekend]

    # Build set of assigned days per doctor
    assigned_days = {doc: set() for doc in doctors}
    for i, doc_name in assignment.items():
        day_idx = duties[i][0]
        assigned_days[doc_name].add(day_idx)

    for doc_name in doctors:
        doc = schedule.doctors[doc_name]
        if doc.fte >= 100:
            continue  # only part‑time

        # # Target final hours based on FTE
        # target_final = (total_initial_hours + total_month_hours) * (doc.fte / 100) / total_fte if total_fte > 0 else 0

        # Instead of duty‑based target, use total weekday hours * FTE fraction
        total_weekdays = len(weekday_indices)
        total_available_hours = total_weekdays * 8.5
        target_final = (doc.fte / 100) * total_available_hours

        current_hours = final_hours.get(doc_name, 0.0)
        excess = current_hours - target_final

        if excess <= 0:
            continue  # no need to suggest

        # Find blank weekdays (no duty, not unavailable, not light‑green)
        blank_days = []
        for day_idx in weekday_indices:
            if (doc_name, day_idx) not in schedule.unavailable and day_idx not in assigned_days[doc_name]:
                # Also skip if light‑green (day‑off preference)
                if hasattr(schedule, 'soft_unavailable') and (doc_name, day_idx) in schedule.soft_unavailable:
                    continue
                blank_days.append(day_idx)

        if not blank_days:
            continue

        # Suggest removing enough blank days to bring excess to zero
        hours_per_day = 8.5
        days_to_remove = int(excess / hours_per_day) + 1
        suggested_days = blank_days[:days_to_remove]

        # Convert day indices to date strings
        suggested_date_strs = [schedule.days[d].date.strftime('%Y-%m-%d') for d in suggested_days]

        rows.append({
            'Doctor': doc_name,
            'FTE (%)': doc.fte,
            'Current Hours': round(current_hours, 1),
            'Target Hours': round(target_final, 1),
            'Excess Hours': round(excess, 1),
            'Suggested Days Off': ', '.join(suggested_date_strs)
        })

    return pd.DataFrame(rows)

def main(template_file=None, output_file=None, config_path='Rules.xlsx', wishes_file=None):
    """
    Run the scheduler with given file paths.
    If a path is None, it will be read from Settings sheet.
    """
    # config_path = 'Rules.xlsx'
    if not os.path.exists(config_path):
        create_default_config(config_path)
        print(f"Created default configuration at {config_path}. Please adjust and re-run.")
        return None

    config = load_config(config_path)
    print(f"[DEBUG] Config loaded from: {config_path}")
    settings = config['Settings'].set_index('Setting')['Value'].to_dict()

    template_file = template_file or settings.get('TemplateFile', '/Users/macjianfeng/Dropbox/github/python/dienstplan/data/external/templates/Stationsplan Oktober 26.xlsx')
    output_file = output_file or settings.get('OutputFile', 'Stationsplan_October_2026_out.xlsx')

    wishes_file = wishes_file or settings.get('WishesFile', None)
    if wishes_file and not os.path.exists(wishes_file):
        print(f"Warning: WishesFile '{wishes_file}' not found. Proceeding without wishes.")
        wishes_file = None
    # Parse
    schedule = parse_template(template_file, config, wishes_path=wishes_file)

    # Always write detected doctors and stations back to Rules.xlsx
    write_missing_config_sheets(schedule, config_path)
    write_skills_auto(schedule, config_path)
    print("Doctors and Stations sheets updated from template.")

    # Reload config: to pick up the newly written sheets
    config = load_config(config_path)   # now the Doctors sheet reflects the template
    print(f"[DEBUG] Config loaded from: {config_path}")
    # Update schedule's duty_types from the reloaded config
    duty_cfg = config.get('DutyTypes', pd.DataFrame())
    if not duty_cfg.empty:
        new_duty_types = {}
        for _, row in duty_cfg.iterrows():
            abbr = str(row['Abbr']).strip()
            fullname = str(row.get('FullName', abbr)).strip()
            requires_senior = str(row.get('RequiresSenior', 'No')).strip().upper() == 'YES'
            weekend_only = str(row.get('WeekendOnly', 'No')).strip().upper() == 'YES'
            priority_val = row.get('Priority', 1)
            if pd.isna(priority_val):
                priority_val = 1
            else:
                priority_val = int(priority_val)
            dt = DutyType(abbr, fullname, requires_senior, weekend_only, priority_val)
            new_duty_types[dt.abbr] = dt
        schedule.duty_types = new_duty_types
    else:
        schedule.duty_types = {
            'SD': DutyType('SD', 'Spätdienst', False, False, 1),
            'ZD': DutyType('ZD', 'Zwischendienst', False, False, 1),
            'KM': DutyType('KM', 'Knochenmarkentnahme', True, False, 2),
        }
    # Ensure KM is present
    if 'KM' not in schedule.duty_types:
        schedule.duty_types['KM'] = DutyType('KM', 'Knochenmarkentnahme', True, False, 2)
        
    # Validate
    errors = validate(schedule, config)
    if errors:
        # Check if the only errors are missing skills and missing demand
        missing_skills = [e for e in errors if 'no skills defined' in e]
        missing_demand = [e for e in errors if 'no duty demand' in e]
        other_errors = [e for e in errors if 'no skills defined' not in e and 'no duty demand' not in e]

        if (missing_skills or missing_demand) and not other_errors:
            print("Configuration incomplete: missing skills and/or station duty counts.")
            print("Generating missing sheets with detected doctors and stations...")
            write_missing_config_sheets(schedule, config_path)
            write_skills_auto(schedule, config_path)
            print("\nPlease open Rules.xlsx and fill in:")
            print("  - Skills sheet: assign each doctor their duty types (SD, ZD, KM, etc.)")
            print("  - Stations sheet: add DutyCounts (e.g., 'SD=2, ZD=1, KM=1')")
            print("Then re-run the script.")
            return
        else:
            print("Validation errors:")
            for e in errors:
                print(f"  - {e}")
            return

    demand = build_demand(schedule, config)
    total_duties = sum(sum(cnts.values()) for cnts in demand.values())
    total_doctor_days = len(schedule.days) * len(schedule.doctors)
    print(f"Total duties required: {total_duties}")
    print(f"Total doctor-days available (before vacations): {total_doctor_days}")
    if total_duties > total_doctor_days:
        print(f"Demand exceeds capacity by {total_duties - total_doctor_days} duties.")
    # Solve
    assignment, solver, duties, doctors, duty_hours, initial_hours = solve_schedule(
        schedule, config, config_path, repair_mode=True
    )

    # After assignment is obtained
    normal_hours = 8.5  # 8:00-16:30
    weekday_indices = [idx for idx, day in enumerate(schedule.days) if not day.is_weekend]

    final_hours = {}
    for doc_name in doctors:
        assigned_hours = 0.0
        assigned_days = set()
        
        # Sum duty hours and collect days with assignments
        for i, assigned_doc in assignment.items():
            if assigned_doc == doc_name:
                day_idx = duties[i][0]
                assigned_hours += duty_hours[i]
                assigned_days.add(day_idx)
        
        # Count normal working weekdays (no duty, not unavailable)
        normal_days = 0
        for day_idx in weekday_indices:
            if (doc_name, day_idx) not in schedule.unavailable and day_idx not in assigned_days:
                normal_days += 1
        
        # Total = initial + duty hours + normal hours
        final_hours[doc_name] = (
            initial_hours.get(doc_name, 0.0)
            + assigned_hours
            + normal_days * normal_hours
        )

    # Generate day-off suggestions
    suggestions_df = generate_day_off_suggestions(
        schedule, assignment, duties, doctors, duty_hours, initial_hours, final_hours
    )
    # Write to WorkingHours sheet
    write_working_hours(config_path, final_hours)
    # Write
    write_output(template_file, output_file, schedule, assignment, duties, doctors, config, solver, suggestions_df)

    print(f"Schedule generated: {output_file}")
    try:
        status = solver.StatusName()
    except (AttributeError, TypeError):
        status = "REPAIR (solution found)"
    print(f"Status: {status}")
    print(f"Objective value: {solver.ObjectiveValue()}")
    try:
        obj = solver.ObjectiveValue()
    except (AttributeError, TypeError):
        obj = "N/A (repair)"
    print(f"Objective value: {obj}")
    try:
        visualize_schedule(schedule, assignment, duties, doctors, output_file.replace('.xlsx', ''))
    except Exception as e:
        print(f"Visualization failed: {e}")
    return output_file
 
def run_scheduler(template_path, output_path, config_path, wishes_path=None, config_dict=None):
    """
    Run the scheduler. If config_dict is provided, it is used instead of reading from config_path.
    """
    log_capture = StringIO()
    sys.stdout = log_capture
    try:
        if config_dict is not None:
            # Write config_dict to a temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                    for sheet, df in config_dict.items():
                        # Ensure DataFrame is clean
                        df_clean = df.fillna("") if isinstance(df, pd.DataFrame) else pd.DataFrame(df)
                        df_clean.to_excel(writer, sheet_name=sheet, index=False)
                temp_path = tmp.name
            main(template_path, output_path, temp_path, wishes_file=wishes_path)
            os.unlink(temp_path)
        else:
            main(template_path, output_path, config_path, wishes_file=wishes_path)
        success = True
    except Exception as e:
        success = False
        raise e
    finally:
        sys.stdout = sys.__stdout__
    return success, log_capture.getvalue()
if __name__ == '__main__':
    main(template_file=None, 
         output_file=None, 
         config_path='/Users/macjianfeng/Dropbox/github/python/dienstplan/Rules.xlsx', 
         wishes_file=None)